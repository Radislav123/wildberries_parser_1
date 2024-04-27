import platform

import openpyxl

from core import models as core_models, parser as parser_core
from core.service import parsing
from . import models, settings


City = dict[str, str]


class Parser(parser_core.Parser):
    settings = settings.Settings()
    parsing_type = core_models.Parsing.Type.POSITION

    def parse_positions(
            self,
            keywords: list[models.Keyword],
            dest: str,
            city: str
    ) -> tuple[list[models.Position], dict[models.Item, Exception]]:
        keywords_dict = {(x.item.vendor_code, x.value): x for x in keywords}
        items_dict = {x.vendor_code: x for x in set(keyword.item for keyword in keywords)}
        positions, errors = parsing.parse_positions(
            [x.item.vendor_code for x in keywords],
            [x.value for x in keywords],
            dest
        )
        errors = {items_dict[vendor_code]: error for vendor_code, error in errors.items()}
        position_objects = [
            models.Position(
                keyword = keywords_dict[key],
                parsing = self.parsing,
                city = city,
                page_capacities = position.page_capacities,
                page = position.page,
                position = position.position,
                promo_page = position.promo_page,
                promo_position = position.promo_position,
                sold_out = position.sold_out
            ) for key, position in positions.items()
        ]

        models.Position.objects.bulk_create(position_objects)

        return position_objects, errors

    @staticmethod
    def get_item_uniq_identifier(item_dict: dict[str, str | int]) -> str:
        return f"{item_dict['vendor_code']}_{item_dict['keyword']}"

    @staticmethod
    def get_keyword_uniq_identifier(keyword: models.Keyword) -> str:
        return f"{keyword.item.vendor_code}_{keyword.value}"

    @classmethod
    def get_position_parser_item_dicts(cls) -> dict[str, dict[str, str | int]]:
        book = openpyxl.load_workbook(cls.settings.PARSER_POSITION_DATA_PATH)
        sheet = book.active
        items = {}
        row = 2
        while sheet.cell(row, 1).value:
            item = {
                "vendor_code": int(sheet.cell(row, 1).value),
                "name": sheet.cell(row, 2).value,
                "keyword": sheet.cell(row, 3).value
            }
            items[cls.get_item_uniq_identifier(item)] = item
            row += 1

        return items

    @classmethod
    def get_position_parser_keywords(cls, divisor: int, remainder: int) -> list[models.Keyword]:
        customer = core_models.ParserUser.get_customer()
        item_dicts = {x[0]: x[1] for index, x in enumerate(cls.get_position_parser_item_dicts().items())
                      if index % divisor == remainder}

        # создание отсутствующих в БД товаров
        old_items_vendor_codes = set(models.Item.objects.values_list("vendor_code", flat = True))
        new_items = {
            key: models.Item(vendor_code = value["vendor_code"], user = customer) for key, value in item_dicts.items()
            if key not in old_items_vendor_codes
        }
        models.Item.objects.bulk_create(new_items.values())
        items = {x.vendor_code: x for x in
                 models.Item.objects.filter(vendor_code__in = [x["vendor_code"] for x in item_dicts.values()])}

        old_keywords_identifiers = set(cls.get_keyword_uniq_identifier(x) for x in models.Keyword.objects.all())
        new_keywords = [
            models.Keyword(
                item = items[value["vendor_code"]],
                value = value["keyword"],
                item_name = value["name"]
            ) for key, value in item_dicts.items() if key not in old_keywords_identifiers
        ]
        models.Keyword.objects.bulk_create(new_keywords)
        keywords = [x for x in models.Keyword.objects.all() if cls.get_keyword_uniq_identifier(x) in item_dicts]
        return keywords

    def run(self, division_remainder: int) -> None:
        # товары пользователей добавляются только при запуске не на машине разработчика
        on_developer_pc = platform.node() == self.settings.secrets.developer.pc_name

        keywords_customer = self.get_position_parser_keywords(
            self.settings.PYTEST_XDIST_WORKER_COUNT,
            division_remainder
        )
        keywords = models.Keyword.objects.filter(id__in = (x.id for x in keywords_customer)).prefetch_related(
            "item",
            "item__user"
        )

        # todo: оставить только Москву - временное решение
        for city_dict in [self.settings.MOSCOW_CITY_DICT]:
            city = city_dict["name"]
            dest = city_dict["dest"]
            _, errors = self.parse_positions(keywords, dest, city)
            self.parsing.not_parsed_items.update(errors)

        if not on_developer_pc:
            keywords_to_prepare = tuple(
                x for x in keywords if x.item not in self.parsing.not_parsed_items
                and x.item.user == core_models.ParserUser.get_customer()
            )
            models.PreparedPosition.prepare(keywords_to_prepare)
