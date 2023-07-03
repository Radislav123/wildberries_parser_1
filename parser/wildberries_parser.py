import logging
import time

import openpyxl
import requests
from requests.exceptions import JSONDecodeError
from selenium.common.exceptions import TimeoutException
from selenium.webdriver import Chrome, ChromeOptions, Remote
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

from logger import Logger
from pages import ItemPage, MainPage, SearchResultsPage
from . import models
from .settings import Settings


City = dict[str, str]
Item = dict[str, str | list[str]]


class LogInException(Exception):
    pass


class WildberriesParser:
    """Отвечает за весь процесс парсинга."""

    settings = Settings()
    logger: logging.Logger
    driver: Chrome
    log_in_driver: Chrome

    def setup_method(self):
        self.logger = Logger("parser")
        self.logger.info("Start")

        options = ChromeOptions()
        # этот параметр тоже нужен, так как в режиме headless с некоторыми элементами нельзя взаимодействовать
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument("--headless")
        options.add_argument("--window-size=1920,1080")
        options.add_experimental_option("excludeSwitches", ["enable-logging"])

        driver_manager = ChromeDriverManager(path = "").install()
        service = Service(executable_path = driver_manager)

        self.driver = Chrome(options = options, service = service)
        self.driver.maximize_window()

    def teardown_method(self):
        self.driver.quit()

    def connect_log_in_driver(self) -> Remote:
        options = ChromeOptions()
        options.add_argument("--headless")
        driver = Remote(self.settings.secrets.wildberries_log_in_driver.url, options = options)
        driver.close()
        driver.session_id = self.settings.secrets.wildberries_log_in_driver.session_id
        return driver

    # не используется, но оставлен
    def find_position_on_page(self, page_number: int, items_number: int, keyword: models.Keyword) -> int:
        """Находит позицию товара на конкретной странице подобно пользователю."""

        search_results_page = SearchResultsPage(self.driver, self.settings, page_number, keyword.value)
        search_results_page.open()
        checked_items = 0
        found = False
        # None возвращаться не должен, так как этот товар точно есть на странице
        position = None

        while not found and items_number > len(search_results_page.items):
            for number, item in enumerate(search_results_page.items[checked_items:], checked_items + 1):
                checked_items += 1
                # ожидание прогрузки
                item.init(item.WaitCondition.VISIBLE)
                item_id = int(item.get_attribute("data-nm-id"))
                if item_id == keyword.item.vendor_code:
                    position = number
                    found = True
                    break
            search_results_page.scroll_down(50)
            search_results_page.items.reset()
        return position

    def find_position(self, city_dict: City, keyword: models.Keyword) -> models.Position:
        """Находит позицию товара в выдаче поиска по ключевому слову среди всех страниц."""

        try:
            page = 1
            position = None
            page_capacities = []
            while page:
                # noinspection SpellCheckingInspection
                url = f"https://search.wb.ru/exactmatch/ru/common/v4/search?appType=1&curr=rub" \
                      f"&dest={city_dict['dest']}&page={page}&query={keyword.value}&regions={city_dict['regions']}" \
                      f"&resultset=catalog&sort=popular&spp=0&suppressSpellcheck=false"
                response = requests.get(url)
                try_number = 0
                try_success = False
                while try_number < self.settings.REQUEST_PAGE_ATTEMPTS_AMOUNT and not try_success:
                    try_number += 1
                    try:
                        page_vendor_codes = [x["id"] for x in response.json()["data"]["products"]]
                        try_success = True
                    except JSONDecodeError:
                        if not try_success and try_number >= self.settings.REQUEST_PAGE_ATTEMPTS_AMOUNT:
                            page = None
                            break
                        else:
                            # еще одна попытка
                            time.sleep(1)
                else:
                    # noinspection PyUnboundLocalVariable
                    page_capacities.append(len(page_vendor_codes))
                    if keyword.item.vendor_code in page_vendor_codes:
                        position = page_vendor_codes.index(keyword.item.vendor_code) + 1
                        break
                    page += 1
        except KeyError as error:
            if "data" in error.args:
                # если возвращаемая позиция == None => товар не был найден по данному ключевому слову
                page_capacities = None
                page = None
                position = None
            else:
                raise error
        return models.Position(
            keyword = keyword,
            city = city_dict["name"],
            page_capacities = page_capacities,
            page = page,
            value = position
        )

    @classmethod
    def get_position_parser_item_dicts(cls) -> list[dict[str, str | int]]:
        book = openpyxl.load_workbook(cls.settings.POSITION_PARSER_DATA_PATH)
        sheet = book.active
        items = []
        row = 2
        while sheet.cell(row, 1).value:
            items.append(
                {
                    "vendor_code": int(sheet.cell(row, 1).value),
                    "name_position": sheet.cell(row, 2).value,
                    "keyword": sheet.cell(row, 3).value
                }
            )
            row += 1
        return items

    @classmethod
    def get_position_parser_keywords(cls) -> list[models.Keyword]:
        item_dicts = cls.get_position_parser_item_dicts()
        # создание отсутствующих товаров в БД
        # noinspection PyStatementEffect
        [models.Item.objects.get_or_create(vendor_code = x["vendor_code"])[0] for x in item_dicts]

        keywords = [models.Keyword.objects.update_or_create(
            value = x["keyword"],
            item_id = x["vendor_code"],
            defaults = {"item_name": x["name_position"]}
        )[0] for x in item_dicts]
        return keywords

    def run_position_parsing(self, city_dict: City) -> None:
        main_page = MainPage(self.driver, self.settings)
        main_page.open()
        dest, regions = main_page.set_city(city_dict["name"])
        city_dict["dest"] = dest
        city_dict["regions"] = regions
        for keyword in self.get_position_parser_keywords():
            position = self.find_position(city_dict, keyword)
            position.save()

    def parse_price(self, item: models.Item) -> tuple[float, float, int | None, int]:
        page = ItemPage(self.driver, self.settings, item.vendor_code)
        page.open()
        page.transfer_cookies(self.log_in_driver)

        try:
            page.sold_out.init_if_necessary()
        except TimeoutException:
            page.price_block.open()
            price = float("".join(page.price_block.price.text.split()[:-1]))
            try:
                final_price = float("".join(page.price_block.final_price.text.split()[:-1]))
                personal_sale = int(page.price_block.personal_sale.text.split()[-1][:-1])
            except TimeoutException:
                final_price = float("".join(page.price_block.price.text.split()[:-1]))
                personal_sale = None
        else:
            price = None
            final_price = None
            personal_sale = None

        reviews_amount = int("".join([x for x in page.review_amount.text.split()[:-1]]))

        return price, final_price, personal_sale, reviews_amount

    @classmethod
    def get_price_parser_items(cls) -> list[models.Item]:
        book = openpyxl.load_workbook(cls.settings.PRICE_PARSER_DATA_PATH)
        sheet = book.active
        items = []
        row = 2
        while sheet.cell(row, 1).value:
            items.append(
                models.Item.objects.update_or_create(
                    vendor_code = sheet.cell(row, 1).value,
                    defaults = {"name_price": sheet.cell(row, 2).value, "category": sheet.cell(row, 3).value}
                )[0]
            )
            row += 1
        return items

    def run_price_parsing(self) -> None:
        self.log_in_driver = self.connect_log_in_driver()

        for item in self.get_price_parser_items():
            price, final_price, personal_sale, reviews_amount = self.parse_price(item)
            price = models.Price(
                item = item,
                price = price,
                final_price = final_price,
                personal_sale = personal_sale,
                reviews_amount = reviews_amount
            )
            price.save()
