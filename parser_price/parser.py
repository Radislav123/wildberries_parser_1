import platform
from typing import Any

import openpyxl

import parser_seller_api.parser
from bot_telegram import bot
from core import models as core_models, parser as parser_core
from core.service import parsing, validators
from parser_price import models, settings


class Parser(parser_core.Parser):
    settings = settings.Settings()
    bot_telegram = bot.Bot()
    parsing_type = core_models.Parsing.Type.PRICE

    def parse_items(
            self,
            items: list[models.Item],
            dest: str
    ) -> tuple[list[models.Price], dict[models.Item, Exception]]:
        items_dict: dict[int, models.Item] = {x.vendor_code: x for x in items
                                              if validators.validate_subscriptions(x.user)}
        items_categories = {x.vendor_code: x.category for x in items_dict.values()}
        prices, errors = parsing.parse_prices(list(items_dict), dest, items_categories)
        errors = {items_dict[vendor_code]: error for vendor_code, error in errors.items()}
        price_objects = []
        for vendor_code, price in prices.items():
            price_object = models.Price(
                item = items_dict[vendor_code],
                parsing = self.parsing,
                reviews_amount = price.reviews_amount,
                price = price.price,
                final_price = price.final_price,
                personal_discount = price.personal_discount,
                sold_out = price.sold_out
            )

            price_objects.append(price_object)
            items_dict[vendor_code].category = price.category
            items_dict[vendor_code].name_site = price.name_site

        models.Price.objects.bulk_create(price_objects)
        models.Item.objects.bulk_update(items_dict.values(), ["category", "name_site"])

        return price_objects, errors

    @classmethod
    def get_price_parser_item_dicts(cls) -> dict[int, dict[str, Any]]:
        book = openpyxl.load_workbook(cls.settings.PARSER_PRICE_DATA_PATH)
        sheet = book.active
        items = {}
        row = 2
        while sheet.cell(row, 1).value:
            item = {
                "vendor_code": sheet.cell(row, 1).value,
                "name": sheet.cell(row, 2).value
            }
            items[item["vendor_code"]] = item
            row += 1
        return items

    @classmethod
    def get_price_parser_items(cls, divisor: int, remainder: int) -> list[models.Item]:
        customer = core_models.ParserUser.get_customer()
        item_dicts = {x[0]: x[1] for index, x in enumerate(cls.get_price_parser_item_dicts().items())
                      if index % divisor == remainder}

        old_items_vendor_codes = set(models.Item.objects.values_list("vendor_code", flat = True))
        new_items = {
            key: models.Item(
                vendor_code = key,
                user = customer,
                name = value["name"]
            ) for key, value in item_dicts.items() if key not in old_items_vendor_codes
        }
        models.Item.objects.bulk_create(new_items.values())
        items = models.Item.objects.filter(vendor_code__in = item_dicts.keys())

        return list(items)

    def run(self) -> None:
        # коммит - feat: парсер скидок теперь запускается только вместе с парсером цен и только в один поток
        items_customer = self.get_price_parser_items(1, 0)
        # товары пользователей добавляются только при запуске не на машине разработчика
        on_developer_pc = platform.node() == self.settings.secrets.developer.pc_name
        if not on_developer_pc:
            items_other = models.Item.objects.exclude(user = core_models.ParserUser.get_customer())
            items = models.Item.objects.filter(
                id__in = (x.id for x in (*items_customer, *items_other))
            ).prefetch_related("user", "category")
        else:
            items = models.Item.objects.filter(
                id__in = (x.id for x in items_customer)
            ).prefetch_related("user", "category")

        parser_api = parser_seller_api.parser.Parser()
        parser_api.setup_method()
        parser_api.run()
        parser_api.teardown_method()

        city_dict = self.settings.MOSCOW_CITY_DICT
        dest = city_dict["dest"]
        prices, errors = self.parse_items(items, dest)
        self.parsing.not_parsed_items = errors

        notifications = models.Price.get_notifications(prices)
        self.bot_telegram.notify(notifications)

        if not on_developer_pc:
            items_to_prepare = tuple(
                x for x in items if x not in errors and x.user == core_models.ParserUser.get_customer()
            )
            models.PreparedPrice.prepare(items_to_prepare)
